"""
Panorama Excel datasource
This datasource doesn't allow field partitions.
It will create a table for each sheet, using the sheet name. Each sheet must have data in a tabular format
Do not leave empty rows or columns.
The first row must have the field names.
"""
import csv
import os

import openpyxl

from panorama_elt.panorama_datalake.panorama_datalake import PanoramaDatalake
from panorama_elt.panorama_logger.setup_logger import log


class XLSDatasource:
    """
    Settings required:
    - table: only one table, corresponding to the file
    - location: path to the local file
    """

    def __init__(
            self,
            datalake: PanoramaDatalake,
            datasource_settings: dict
    ):

        self.table_fields = {}
        table_settings = datasource_settings.get('tables')
        if table_settings:
            for table_setting in table_settings:
                fields = table_setting.get('fields')
                if fields:
                    self.table_fields[table_setting.get('name')] = [f.get("name") for f in fields]

        self.location = datasource_settings.get('location')
        self.datalake = datalake

    def test_connections(self) -> dict:
        """
        Performs connections test
        :return: dict with test results
        """
        from pathlib import Path
        path = Path(self.location)

        results = {'XLS': 'OK' if path.is_file() else 'File {} not found'.format(self.location)}

        return results

    def get_tables(self) -> list:
        """
        Returns the list of sheet names, as a list of tables
        :return: list of sheet names
        """
        workbook = openpyxl.load_workbook(self.location)
        return workbook.get_sheet_names()

    def get_fields(self, table: str, force_query: bool = False) -> list:
        """
        Returns a list of fields of the table based on the first row of the specified sheet in the Excel file.
        All types are assumed to be string.

        :param table: table name
        :param force_query: (optional) if set to True, will query the db even if there is a definition set
        :return: list[str] of fields
        """

        # If the field list is declared in the settings file, return it.
        if self.table_fields and self.table_fields.get(table) and not force_query:
            return self.table_fields.get(table)

        workbook = openpyxl.load_workbook(self.location)
        sheet = workbook.get_sheet_by_name(table)
        fields = []

        colnum = 1
        value = sheet.cell(row=1, column=1).value
        while value:
            fields.append(value)
            colnum += 1
            value = sheet.cell(row=1, column=colnum).value

        workbook.close()

        log.debug("Fields in table: {}".format(fields))

        fields_list = []
        for field in fields:
            fields_list.append({"name": field, "type": 'string'})

        return fields_list

    def extract_and_load(self, selected_tables: str = None, force: bool = False):
        """
        Upload the file to the datalake

        :param selected_tables: (optional) list of tables to extract and load
        :param force: Forces a full update of all the partitions
        :return:
        """

        workbook = openpyxl.load_workbook(self.location)
        for table, fields in self.table_fields.items():
            sheet = workbook.get_sheet_by_name(table)

            rownum = 2
            dataset = []
            while rownum < 1000000:
                row = []
                for colnum in range(len(fields)):
                    row.append(sheet.cell(row=rownum, column=colnum+1).value)
                rownum += 1
                if all([v is None for v in row]):
                    break
                dataset.append(row)

            # Save the dataset in a csv file
            filename = "{}.csv".format(table)
            with open(filename, 'w') as f:
                write = csv.writer(f, doublequote=False, escapechar='\\')
                write.writerow(fields)
                write.writerows(dataset)

            self.datalake.upload_table_from_file(filename=filename, table=table, update_partitions=True)

            os.remove(filename)

        workbook.close()
